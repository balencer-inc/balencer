#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""月次の請求書を一括生成する（HTML→PDF＋売掛管理表/数字マスターの貼り付け行＋検算）。

    python3 .claude/skills/invoice/scripts/invoice_build.py docs/invoices/2026-08/invoices.json

入力JSON:
{
  "issue_month": "2026-08",                    # 請求月。Drive の「2026.08請求書」に対応
  "invoices": [
    {"client": "hilltop",                      # docs/company/invoice-clients.json のキー
     "no": "07-080", "date": "2026-09-01",
     "due": null,                              # 省略時は client の due_rule から計算
     "expected_total_incl": 33000,             # ①売掛金一覧の「本体税込」。突合用
     "items": [{"name": "サーバー維持管理費", "note": "", "detail": "8月分",
                "amount": 33000, "basis": "incl", "rate": 10}],
     "remarks": ["契約に基づく…"],             # 振込手数料の一文は自動で最後に付く
     "plan_items": [{"case": "サーバー保守 月額", "excl": 30000}],   # ④用。省略可
     "besshi": {"rows": [...], "notes": [...]}                      # 別紙。省略可
    }
  ]
}

出力（docs/invoices/<issue_month>/ 配下）:
  <No>_<short>.html / YYYYMMDD【<short>様】.pdf   … PDFはDriveの命名規則そのまま
  売掛管理表_貼り付け.tsv                          … ③「売上入力」シートへ
  数字マスター_売上案件別.tsv                      … ④「売上_案件別」へ
"""
import calendar
import json
import os
import re
import subprocess
import sys
from datetime import date, timedelta
from math import floor

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", ".."))
TEMPLATE = os.path.join(ROOT, ".claude/skills/invoice/assets/invoice-template.html")
CLIENTS = os.path.join(ROOT, "docs/company/invoice-clients.json")
SEAL = os.path.join(ROOT, "docs/invoices/99_local/seal-balencer.png")
BANK_FILE = os.path.join(ROOT, "docs/invoices/99_local/振込先.md")
LEDGER = os.path.join(ROOT, "docs/invoices/ledger.tsv")
# 採番の正本＝経理の「請求書ナンバー管理表」。Googleドライブのデスクトップ同期で
# ローカルに見えているので openpyxl で直接読める。無ければ ledger.tsv にフォールバック。
NO_MASTER = os.path.expanduser(
    "~/Library/CloudStorage/GoogleDrive-tabe@balencer.jp/マイドライブ/"
    "バレンサー経理用/請求書/請求書ナンバー管理表.xlsx")
CHROME = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
FEE_NOTE = "恐れ入りますが振込手数料は貴社にてご負担をお願いいたします。"


def ymd(s):
    y, m, d = (int(x) for x in s.split("-"))
    return date(y, m, d)


def month_end(y, m):
    return date(y, m, calendar.monthrange(y, m)[1])


def add_months(y, m, n):
    t = (y * 12 + m - 1) + n
    return t // 12, t % 12 + 1


def calc_due(issue: date, period: tuple, rule: str):
    """支払期日。締め月(period)起点の規則と、請求日(issue)起点の規則がある。

    ① 売掛金一覧の「入金予定日」が唯一の正。ここは入力漏れ時のフォールバック。
    """
    py, pm = period
    if rule == "same_month_end":
        return month_end(py, pm)
    if rule == "next_month_end":
        return month_end(*add_months(py, pm, 1))
    if rule == "next_month_10":
        y, m = add_months(py, pm, 1)
        return date(y, m, 10)
    if rule == "next2_month_10":
        y, m = add_months(py, pm, 2)
        return date(y, m, 10)
    if rule == "issue_plus_30":
        return issue + timedelta(days=30)
    if rule == "issue_plus_60":
        return issue + timedelta(days=60)
    return None


def jp(d: date) -> str:
    return f"{d.year}年{d.month}月{d.day}日"


def to_excl(amount: int, basis: str, rate: int) -> int:
    if basis == "excl":
        return int(amount)
    return int(round(amount / (1 + rate / 100)))


def fiscal_term(d: date) -> int:
    """期。12月始まり（第7期＝2025/12〜2026/11）。"""
    return d.year - 2018 if d.month == 12 else d.year - 2019


def read_no_master(term: int):
    """請求書ナンバー管理表から、その期の使用済み番号を読む。

    戻り値: {No: (社名, 作成日, 請求金額)}。読めなければ None。
    """
    if not os.path.exists(NO_MASTER):
        return None
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(NO_MASTER, read_only=True, data_only=True)
    except Exception:
        return None
    z = str.maketrans("０１２３４５６７８９", "0123456789")
    sheet = next((n for n in wb.sheetnames
                  if n.translate(z).strip().startswith(f"{term}期")), None)
    if sheet is None:
        return None
    out = {}
    for no, name, made, amount in wb[sheet].iter_rows(min_row=2, max_col=4, values_only=True):
        if no and name:
            out[str(no).strip()] = (str(name).strip(), made, amount)
    return out


def read_ledger():
    rows = []
    if not os.path.exists(LEDGER):
        return rows
    for line in open(LEDGER, encoding="utf-8"):
        line = line.rstrip("\n")
        if not line or line.startswith("#") or line.startswith("No\t"):
            continue
        rows.append(line.split("\t"))
    return rows


def next_no(used: set, d: date) -> str:
    """その期で未使用の最小番号を返す（自動採番）。"""
    term = f"{fiscal_term(d):02d}"
    n = max((int(x.split("-")[1]) for x in used if x.startswith(term + "-")), default=0)
    while f"{term}-{n + 1:03d}" in used:
        n += 1
    return f"{term}-{n + 1:03d}"


def append_ledger(rows):
    with open(LEDGER, "a", encoding="utf-8") as f:
        for r in rows:
            f.write("\t".join(str(x) for x in r) + "\n")


def bank_account() -> str:
    """口座番号は git 管理外の 99_local/振込先.md から読む。無ければ伏せ字。"""
    try:
        txt = open(BANK_FILE, encoding="utf-8").read()
    except OSError:
        return "普通　（99_local/振込先.md 未設置）"
    m = re.search(r"(普通|当座)[^\d]{0,8}(\d{6,8})", txt)
    return f"{m.group(1)}　{m.group(2)}" if m else "普通　（振込先.md から読めず）"


def totals(items):
    sub = {10: 0, 8: 0}
    for it in items:
        rate = int(it.get("rate", 10))
        sub[rate] += to_excl(int(it["amount"]), it.get("basis", "excl"), rate)
    tax = {r: floor(sub[r] * r / 100) for r in sub}
    return sub, tax, sum(sub.values()) + sum(tax.values())


def build_besshi(html: str, besshi: dict) -> str:
    """別紙の明細行を組み、月ごとの小計と全体合計を自動で付ける。

    行: {"date": "7/25", "place": ..., "person": ..., "allowance": 10000,
         "route": ..., "fare": 860}　※金額はすべて税込の数値
    """
    out, group, gsum = [], None, None

    def flush():
        if group is None:
            return
        out.append(f'    <tr class="st"><td colspan="3">{group}月分 小計</td>'
                   f'<td class="c">¥{gsum[0]:,}</td><td></td>'
                   f'<td class="r">¥{gsum[1]:,}</td></tr>')

    tot = [0, 0]
    for r in besshi["rows"]:
        mon = str(r["date"]).split("/")[0]
        if mon != group:
            flush()
            group, gsum = mon, [0, 0]
        a, f = int(r.get("allowance", 0)), int(r.get("fare", 0))
        gsum[0] += a
        gsum[1] += f
        tot[0] += a
        tot[1] += f
        out.append(f'    <tr><td>{r["date"]}</td><td class="nm">{r["place"]}</td>'
                   f'<td class="nm">{r["person"]}</td><td class="c">¥{a:,}</td>'
                   f'<td class="nm">{r["route"]}</td><td class="r">¥{f:,}</td></tr>')
    flush()
    out.append(f'    <tr class="gr"><td colspan="3">合計（税込）</td>'
               f'<td class="c">¥{tot[0]:,}</td><td></td><td class="r">¥{tot[1]:,}</td></tr>')

    html = re.sub(r'    <!-- 別紙行 .*?class="gr">.*?</tr>\n', "\n".join(out) + "\n",
                  html, flags=re.S)
    bn = "\n".join(f"    <li>{n}</li>" for n in besshi.get("notes", []))
    return re.sub(r'    <li>\{\{BESSHI_NOTE_1\}\}</li>\n    <li>\{\{BESSHI_NOTE_2\}\}</li>',
                  bn, html)


def render(inv, cl, sub, tax, total, issue: date, due: date, out_html: str):
    html = open(TEMPLATE, encoding="utf-8").read()

    rows = []
    for it in inv["items"]:
        rate = int(it.get("rate", 10))
        excl = to_excl(int(it["amount"]), it.get("basis", "excl"), rate)
        note = it.get("note", "")
        nm = f'{it["name"]}<span class="d">　{note}</span>' if note else it["name"]
        mark = "※" if rate == 8 else ""
        rows.append(f'  <tr><td>{nm}</td><td>{it.get("detail","")}</td>'
                    f'<td class="r">{mark}¥{excl:,}</td></tr>')
    html = re.sub(r'  <tr><td>\{\{ITEM_NAME\}\}.*?</tr>', "\n".join(rows), html, flags=re.S)

    remarks = list(inv.get("remarks", [])) + [FEE_NOTE]
    li = "\n".join(f"    <li>{r}</li>" for r in remarks)
    html = re.sub(r'    <li>\{\{REMARK_CONTRACT\}\}.*?\{\{REMARK_TAX\}\}[^\n]*\n', li + "\n",
                  html, flags=re.S)

    besshi = inv.get("besshi")
    if besshi:
        html = build_besshi(html, besshi)
    else:
        i = html.index("<!-- ▼ 別紙ページ")
        html = html[:i] + "</body></html>"

    seal_rel = os.path.relpath(SEAL, os.path.dirname(os.path.abspath(out_html)))
    for k, v in {
        "{{TO_NAME}}": cl["name"], "{{TO_ADDRESS}}": inv.get("to_address", ""),
        "{{ISSUE_DATE}}": jp(issue), "{{INVOICE_NO}}": inv["no"], "{{SEAL_PATH}}": seal_rel,
        "{{TOTAL_INCL}}": f"{total:,}", "{{DUE_DATE}}": jp(due),
        "{{BANK_ACCOUNT}}": bank_account(),
        "{{SUB10}}": f"{sub[10]:,}", "{{SUB8}}": f"{sub[8]:,}",
        "{{TAX10}}": f"{tax[10]:,}", "{{TAX8}}": f"{tax[8]:,}",
    }.items():
        html = html.replace(k, v)

    html = re.sub(r"[ \t]*<!--\s*(明細行|別紙行|▼ 別紙ページ).*?-->\n", "", html, flags=re.S)
    left = sorted(set(re.findall(r"\{\{[A-Z_0-9]+\}\}", html)))
    if left:
        raise SystemExit(f"未展開のプレースホルダ {left} ({inv['no']})")
    with open(out_html, "w", encoding="utf-8") as f:
        f.write(html)


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 2
    data = json.load(open(sys.argv[1], encoding="utf-8"))
    master = json.load(open(CLIENTS, encoding="utf-8"))["clients"]
    month = data["issue_month"]
    py, pm = (int(x) for x in month.split("-")[:2])
    outdir = os.path.join(ROOT, "docs/invoices", month)
    os.makedirs(outdir, exist_ok=True)

    ledger = read_ledger()
    used = {r[0] for r in ledger}
    no_master = read_no_master(fiscal_term(date(py, pm, 1)))
    if no_master is None:
        print("⚠ 請求書ナンバー管理表が読めません。ledger.tsv だけで採番します")
    else:
        print(f"採番の正本: 請求書ナンバー管理表（{len(master)}件 使用済み・"
              f"最終 {max(no_master)}）")
        used |= set(no_master)
    new_ledger = []
    ar_rows, plan_rows, pdfs, ng, warn_due, auto, conflict = [], [], [], [], [], [], []
    accepted = []
    print("=" * 74)
    print(f"請求書 一括生成　{month}　{len(data['invoices'])}件")
    print("=" * 74)

    for inv in data["invoices"]:
        cl = master.get(inv["client"])
        if not cl:
            raise SystemExit(f"請求先マスタに無いキー: {inv['client']}")
        issue = ymd(inv["date"])
        if not inv.get("no"):
            inv["no"] = next_no(used, issue)
            auto.append((inv["no"], cl["name"]))
        used.add(inv["no"])
        if no_master and inv["no"] in no_master:
            mname = no_master[inv["no"]][0]
            names = [cl["name"], cl["short"], cl.get("ledger_alias") or ""]
            if not any(n and (mname in n or n in mname) for n in names):
                (accepted if inv.get("no_conflict_ok") else conflict).append(
                    (inv["no"], cl["name"], mname, inv.get("no_conflict_ok", "")))
        if inv.get("due"):
            due, guessed = ymd(inv["due"]), ""
        else:
            due, guessed = calc_due(issue, (py, pm), cl["due_rule"]), "（推定）"
            if due is None:
                raise SystemExit(f"{inv['no']}: due_rule が manual。due を明示してください")
            warn_due.append((inv["no"], cl["name"], due))

        sub, tax, total = totals(inv["items"])
        exp = inv.get("expected_total_incl")
        mark = "  " if exp is None or total == int(exp) else "NG"
        if mark == "NG":
            ng.append((inv["no"], cl["name"], total, int(exp)))

        new_ledger.append([inv["no"], cl["name"], month, f"{issue:%Y-%m-%d}", total,
                           "invoice_build"])
        base = f'{inv["no"]}_{cl["short"]}'
        html_path = os.path.join(outdir, base + ".html")
        pdf_name = f'{issue:%Y%m%d}【{cl["short"]}様】.pdf'
        render(inv, cl, sub, tax, total, issue, due, html_path)
        pdfs.append((html_path, os.path.join(outdir, pdf_name)))

        print(f'{mark} {inv["no"]}  {cl["name"]:<22} 税抜{sum(sub.values()):>10,} '
              f'税{sum(tax.values()):>8,}  税込{total:>10,}  期日 {due:%Y/%m/%d}{guessed}')
        for it in inv["items"]:
            r = int(it.get("rate", 10))
            print(f'      ・{it["name"]}　{it.get("detail","")}　'
                  f'{to_excl(int(it["amount"]), it.get("basis","excl"), r):,}（{r}%）')

        # ③ 売掛金管理表「売上入力」の貼り付け行
        品名 = "、".join(dict.fromkeys(i["name"] for i in inv["items"]))
        ar_rows.append([
            issue.year - 2018, issue.month, issue.day, cl["ar_code"] or "要採番",
            cl["ar_name"], 20 if cl["due_rule"] == "next2_month_10" else 31,
            品名, 1, total, total, pm, "",
            due.year - 2018, due.month, due.day, inv["no"], pm, total,
        ])
        # ④ 営業管理数字マスター「売上_案件別」（税抜）
        for p in inv.get("plan_items") or [{"case": 品名, "excl": sum(sub.values())}]:
            plan_rows.append([cl["plan_name"], p["case"], "A 確定",
                              f"{py}-{pm:02d}", p["excl"]])

    # PDF化
    print("-" * 74)
    for html_path, pdf_path in pdfs:
        subprocess.run([CHROME, "--headless", f"--print-to-pdf={pdf_path}",
                        "--no-pdf-header-footer", html_path],
                       capture_output=True, check=False)
        ok = os.path.exists(pdf_path)
        print(("PDF " if ok else "失敗 ") + os.path.basename(pdf_path))

    ar_head = ["年(和暦)", "月", "日", "取引先コード", "取引先名", "締日", "品名", "数量",
               "単価(税込)", "売上金額(税込)", "請求月(L列)", "受入金額",
               "入金年(和暦)", "入金月", "入金日", "請求書No", "請求月(O列)", "請求金額(税込)"]
    p1 = os.path.join(outdir, "売掛管理表_貼り付け.tsv")
    with open(p1, "w", encoding="utf-8") as f:
        f.write("\t".join(ar_head) + "\n")
        for r in ar_rows:
            f.write("\t".join(str(x) for x in r) + "\n")
    p2 = os.path.join(outdir, "数字マスター_売上案件別.tsv")
    with open(p2, "w", encoding="utf-8") as f:
        f.write("顧客\t案件名\t確度\t対象月\t金額(税抜)\n")
        for r in plan_rows:
            f.write("\t".join(str(x) for x in r) + "\n")

    fresh = [r for r in new_ledger if r[0] not in {x[0] for x in ledger}]
    if fresh:
        append_ledger(fresh)
    print("-" * 74)
    if auto:
        print("自動採番:")
        for no, nm in auto:
            print(f"   {no}  {nm}")
    print(f"出力先 {os.path.relpath(outdir, ROOT)}")
    print(f"  ③貼り付け {os.path.basename(p1)}　④確認 {os.path.basename(p2)}")
    if warn_due:
        print("\n⚠ 入金予定日が入力に無く、契約条件から推定しました（①で要確認）:")
        for no, nm, d in warn_due:
            print(f"   {no} {nm}: {d:%Y/%m/%d}")
    for no, mine, theirs, why in accepted:
        print(f"\n（了承済みの不一致）{no}  請求書『{mine}』／管理表『{theirs}』… {why}")
    if conflict:
        print("\n■ 請求書ナンバー管理表と宛先が食い違っています（正本が優先）:")
        for no, mine, theirs, _ in conflict:
            print(f"   {no}  この請求書『{mine}』 ／ 管理表『{theirs}』")
    if ng:
        print("\n⚠ 売掛金一覧(①)の税込と合わない請求があります:")
        for no, nm, got, exp in ng:
            print(f"   {no} {nm}: 生成 {got:,} / ①は {exp:,}（差 {got-exp:+,}）")
        return 1
    print("\n① 売掛金一覧の『本体税込』と全件一致")
    return 1 if conflict else 0


if __name__ == "__main__":
    sys.exit(main())
