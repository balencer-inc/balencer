#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""請求書のExcel版を作る（経理がExcelで触ってPDFにする用）。

    python3 .claude/skills/invoice/scripts/invoice_xlsx.py docs/invoices/2026-08/invoices.json

invoice_build.py と同じ入力JSONから、同じ書式の .xlsx を1請求1ファイル出力する。
HTML版と違うのは「触れること」。集計は数式で入っているので、明細を書き換えれば
小計・消費税・合計が自動で追随する。A4縦の印刷範囲・社印つき。

  ・小計    = SUM(明細範囲)
  ・消費税  = ROUNDDOWN(小計 * 税率)   ※税率ごとに1回だけ（インボイス制度）
  ・合計    = 小計 + 消費税
"""
import json
import os
import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from invoice_build import (  # noqa: E402
    CLIENTS, ROOT, SEAL, bank_account, calc_due, jp, to_excl, ymd,
)

NAVY = "22364F"
GREY = "6B7789"
LINE = "D3DAE2"
SOFT = "F5F8FA"
FONT = "Yu Gothic"
ROWS_MIN = 8          # 明細の行数（足りなければ増やす）


def F(sz=10, b=False, color="1A2230"):
    return Font(name=FONT, size=sz, bold=b, color=color)


def side(style="thin", color=LINE):
    return Side(style=style, color=color)


def build(inv, cl, issue: date, due: date, out: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "請求書"
    ws.sheet_view.showGridLines = False

    widths = [17, 17, 12, 11, 13, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def put(cell, value, font=None, align=None, fill=None, border=None):
        c = ws[cell]
        c.value = value
        c.font = font or F()
        if align:
            c.alignment = align
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        if border:
            c.border = border
        return c

    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 見出し
    ws.merge_cells("A1:F1")
    put("A1", "請　求　書", F(20, True), center)
    ws.row_dimensions[1].height = 38

    # 宛名（左）
    ws.merge_cells("A3:C3")
    a = put("A3", cl["name"], F(14), Alignment(horizontal="left", vertical="bottom"))
    for col in "ABC":
        ws[f"{col}3"].border = Border(bottom=side("thin", "1A2230"))
    ws.row_dimensions[3].height = 24
    put("A4", "御中", F(11))
    if inv.get("to_address"):
        ws.merge_cells("A5:C6")
        put("A5", inv["to_address"].replace("<br>", "\n"), F(8.5, color=GREY), left)

    # 発行者（右）
    put("E3", "作成日：", F(9, color=GREY), right)
    put("F3", f"{issue:%Y/%m/%d}", F(9), right)
    put("E4", "No：", F(9, color=GREY), right)
    put("F4", inv["no"], F(9), right)
    issuer = ["株式会社バレンサー", "登録番号：T8160001022136", "〒530-0001",
              "大阪府大阪市北区梅田1-11-4", "大阪駅前第4ビル9階 923 1542号",
              "連絡先：06-4400-5365"]
    for i, line in enumerate(issuer):
        ws.merge_cells(f"E{6 + i}:F{6 + i}")
        put(f"E{6 + i}", line, F(12, True) if i == 0 else F(8.5),
            Alignment(horizontal="left", vertical="center"))

    if os.path.exists(SEAL):
        img = XLImage(SEAL)
        img.width = img.height = 74          # 約21mm
        ws.add_image(img, "F8")

    put("A13", "下記の通りご請求申し上げます。", F(10))

    # 合計金額
    ws.merge_cells("A15:C15")
    ws.merge_cells("D15:F15")
    box = Border(left=side("medium", NAVY), right=side("medium", NAVY),
                 top=side("medium", NAVY), bottom=side("medium", NAVY))
    put("A15", "合計金額", F(12), Alignment(horizontal="left", vertical="center"), SOFT, box)
    put("D15", None, F(18, True), right, SOFT, box)
    for col in "BCEF":
        ws[f"{col}15"].border = box
        ws[f"{col}15"].fill = PatternFill("solid", fgColor=SOFT)
    ws.row_dimensions[15].height = 34

    # 明細
    head = 17
    for col, label, w in ((1, "商品", None), (4, "詳細", None), (5, "金額", None)):
        pass
    ws.merge_cells(f"A{head}:C{head}")
    ws.merge_cells(f"E{head}:F{head}")
    put(f"A{head}", "　商品", F(9, True, "FFFFFF"), left, NAVY)
    put(f"D{head}", "詳細", F(9, True, "FFFFFF"), center, NAVY)
    put(f"E{head}", "金額　", F(9, True, "FFFFFF"), right, NAVY)
    for col in "BCF":
        ws[f"{col}{head}"].fill = PatternFill("solid", fgColor=NAVY)

    items = inv["items"]
    n = max(len(items), ROWS_MIN)
    first, last = head + 1, head + n
    under = Border(bottom=side())
    for i in range(n):
        r = first + i
        ws.merge_cells(f"A{r}:C{r}")
        ws.merge_cells(f"E{r}:F{r}")
        it = items[i] if i < len(items) else None
        name = ""
        if it:
            name = it["name"] + (f"　{it['note']}" if it.get("note") else "")
        put(f"A{r}", name or None, F(9.5), left, border=under)
        put(f"D{r}", (it or {}).get("detail"), F(9.5), center, border=under)
        v = None
        if it:
            rate = int(it.get("rate", 10))
            v = to_excl(int(it["amount"]), it.get("basis", "excl"), rate)
        c = put(f"E{r}", v, F(9.5), right, border=under)
        c.number_format = '"¥"#,##0'
        for col in "BCF":
            ws[f"{col}{r}"].border = under
        # 8%対象は D 列に印
        if it and int(it.get("rate", 10)) == 8:
            ws[f"D{r}"].value = f"{it.get('detail','')} ※"

    # 集計（数式）
    rows8 = [first + i for i, it in enumerate(items) if int(it.get("rate", 10)) == 8]
    rng10 = f"E{first}:E{last}"
    sum8 = "+".join(f"E{r}" for r in rows8) if rows8 else "0"
    s = last + 2
    labels = [("小計（10%対象）", f"=SUM({rng10})-({sum8})"),
              ("　　　（8%対象）", f"={sum8}"),
              ("消費税（10%対象）", f"=ROUNDDOWN(E{s}*0.1,0)"),
              ("　　　　　（8%対象）", f"=ROUNDDOWN(E{s + 1}*0.08,0)"),
              ("合計金額", f"=E{s}+E{s + 1}+E{s + 2}+E{s + 3}")]
    for i, (label, formula) in enumerate(labels):
        r = s + i
        ws.merge_cells(f"C{r}:D{r}")
        put(f"C{r}", label, F(9.5, color=GREY if i < 4 else "1A2230"), right)
        ws.merge_cells(f"E{r}:F{r}")
        c = put(f"E{r}", formula, F(9.5), right)
        c.number_format = '"¥"#,##0'
        for col in "CDEF":
            ws[f"{col}{r}"].border = Border(bottom=side("thin", "EEF1F5"))
    fin = s + 4
    for col in "CDEF":
        ws[f"{col}{fin}"].border = Border(top=side("medium", NAVY), bottom=side("double", NAVY))
        ws[f"{col}{fin}"].fill = PatternFill("solid", fgColor=SOFT)
    ws[f"C{fin}"].font = F(11.5, True)
    ws[f"E{fin}"].font = F(11.5, True)
    ws[f"D15"] = f"=E{fin}"
    ws["D15"].number_format = '"¥"#,##0"（税込）"'

    # 振込先・支払期日
    b = fin + 2
    box2 = Border(left=side(), right=side(), top=side(), bottom=side())
    ws.merge_cells(f"A{b}:C{b}")
    put(f"A{b}", "【振込先】", F(9, True, "3D5A80"), left)
    for i, (k, v) in enumerate([("金融機関", "みずほ銀行　大津支店"),
                                ("口座", bank_account()),
                                ("口座名義", "株式会社バレンサー")]):
        put(f"A{b + 1 + i}", k, F(9, color=GREY))
        ws.merge_cells(f"B{b + 1 + i}:C{b + 1 + i}")
        put(f"B{b + 1 + i}", v, F(9.5), left)
    ws.merge_cells(f"E{b}:F{b}")
    put(f"E{b}", "お支払い期日", F(9, True, "3D5A80"), center, SOFT)
    ws.merge_cells(f"E{b + 1}:F{b + 2}")
    put(f"E{b + 1}", f"{due:%Y/%m/%d}", F(13, True), center, SOFT)
    for r in range(b, b + 3):
        for col in "EF":
            ws[f"{col}{r}"].border = box2

    # 備考
    k = b + 5
    put(f"A{k}", "備考", F(9, True, "3D5A80"))
    ws[f"A{k}"].border = Border(bottom=side())
    for col in "BCDEF":
        ws[f"{col}{k}"].border = Border(bottom=side())
    remarks = list(inv.get("remarks", [])) + [
        "恐れ入りますが振込手数料は貴社にてご負担をお願いいたします。"]
    for i, t in enumerate(remarks):
        r = k + 1 + i
        ws.merge_cells(f"A{r}:F{r}")
        put(f"A{r}", "・" + t, F(8.5, color="48566B"), left)
        ws.row_dimensions[r].height = 13 + 11 * (len(t) // 60)

    last_row = k + len(remarks)
    ws.print_area = f"A1:F{last_row}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.5
    ws.page_margins.top = ws.page_margins.bottom = 0.5

    wb.save(out)


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 2
    data = json.load(open(sys.argv[1], encoding="utf-8"))
    clients = json.load(open(CLIENTS, encoding="utf-8"))["clients"]
    month = data["issue_month"]
    py, pm = (int(x) for x in month.split("-")[:2])
    outdir = os.path.join(ROOT, "docs/invoices", month)
    os.makedirs(outdir, exist_ok=True)

    for inv in data["invoices"]:
        cl = clients[inv["client"]]
        issue = ymd(inv["date"])
        due = ymd(inv["due"]) if inv.get("due") else calc_due(issue, (py, pm), cl["due_rule"])
        out = os.path.join(outdir, f'{inv["no"]}_{cl["short"]}.xlsx')
        build(inv, cl, issue, due, out)
        print(f'Excel {os.path.basename(out)}')
    print(f"\n出力先 {os.path.relpath(outdir, ROOT)}")
    print("集計は数式。明細を書き換えれば小計・消費税・合計が追随する")
    return 0


if __name__ == "__main__":
    sys.exit(main())
